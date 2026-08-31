from __future__ import annotations

import math
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

from pareto_designer.views.experiment_report.config import (
    ConfigError,
    load_experiment_config,
)
from pareto_designer.views.experiment_report.excel_tables import (
    DESIGN_RUN_HEADERS,
    ALPHA_CORREL_SECTIONS,
    design_run_row,
    write_correlations_block,
    write_data_block,
)
from pareto_designer.views.experiment_report.metrics import (
    aggregate_cross_sequence,
    build_design_run_summaries,
    SWEEP_CORREL_METRICS,
)
from pareto_designer.models.context import ParetoResult
from pareto_designer.views.experiment_report.models import (
    LoadedRun,
    RunParams,
    SamplerParams,
)
from pareto_designer.views.experiment_report.paths import (
    db_fsm_state_count_for_motif,
    fsm_size_from_id,
    parse_run_dir,
)
from pareto_designer.views.experiment_report.sweeps import (
    alpha_regime,
    sweep_membership,
)
from pareto_designer.shared.seq_design_utils.run_grid import GridMode, run_design_grid

CONFIG_PATH = (
    Path(__file__).resolve().parents[2] / "configs" / "pareto_experiment_ma0267.yaml"
)


def _make_result(cost: float, binding: float, origin: float, sid: str = "001"):
    return ParetoResult(
        cost=cost,
        binding_score=binding,
        origin_binding_score=origin,
        id=sid,
        url=f"{sid}_details.html",
        txt_file=f"{sid}.txt",
        fasta_file=f"{sid}.fa",
        positional_objectives_file=f"{sid}.npy",
        max_positional_cost=1.0,
        min_positional_binding=0.0,
        max_positional_binding=1.0,
        sequence="ACGT",
        n_cost_items=1,
        motif_hits=[],
    )


def _make_run(
    seq_id: str,
    k: int,
    alpha: float,
    log_pos: bool,
    reduce_fsm_by: float,
    fsm_size: int,
    solutions: list[ParetoResult],
) -> LoadedRun:
    fsm_id = "logexp_db_fsm" if reduce_fsm_by == 0 else f"logexp_reduced_fsm_{fsm_size}"
    params = RunParams(
        seq_id=seq_id,
        fsm_id=fsm_id,
        fsm_size=fsm_size,
        reduce_fsm_by=reduce_fsm_by,
        sampler=SamplerParams(k=k, alpha=alpha, log_pos=log_pos),
    )
    return LoadedRun(
        params=params,
        metadata={
            "n_solutions": len(solutions),
            "runtime_seconds": 1.0,
            "fsm_binding_score_err": 0.1,
        },
        solutions=solutions,
        path=Path("dummy/results_metadata.json"),
    )


def test_load_valid_config():
    config = load_experiment_config(CONFIG_PATH)
    assert config.name == "pareto_parameter_sweep_ma0267"
    assert len(config.expected_runs()) == 13 * len(config.seq_files())


def test_reject_unknown_top_level_key(tmp_path: Path):
    bad = tmp_path / "bad.yaml"
    bad.write_text("name: x\nfixed: {}\nextra: 1\nsweeps: {}\n", encoding="utf-8")
    with pytest.raises(ConfigError, match="unknown keys"):
        load_experiment_config(bad)


def test_reject_invalid_sampler_alpha(tmp_path: Path):
    text = CONFIG_PATH.read_text(encoding="utf-8").replace(
        "0.0, 0.5, 1.0", "bad_alpha, 0.5, 1.0"
    )
    bad = tmp_path / "bad.yaml"
    bad.write_text(text, encoding="utf-8")
    with pytest.raises(ConfigError, match="invalid sampler_alpha"):
        load_experiment_config(bad)


def test_fsm_size_from_id():
    db_size = db_fsm_state_count_for_motif("MA0267.1")
    assert fsm_size_from_id("logexp_db_fsm", db_fsm_size=db_size) == (db_size, 0.0)
    assert fsm_size_from_id("logexp_reduced_fsm_2048", db_fsm_size=db_size)[0] == 2048


def test_parse_run_dir(tmp_path: Path):
    run_dir = (
        tmp_path
        / "seq_a"
        / "transition_0.50"
        / "MA0267.1"
        / "logexp_reduced_fsm_2048"
        / "PowerLawSUS"
        / "k_100__alpha_1.0_log_pos"
    )
    run_dir.mkdir(parents=True)
    params = parse_run_dir(run_dir)
    assert params.seq_id == "seq_a"
    assert params.sampler.k == 100
    assert params.sampler.alpha == 1.0
    assert params.sampler.log_pos is True


def test_alpha_regime():
    params = RunParams(
        seq_id="s",
        fsm_id="logexp_reduced_fsm_2048",
        fsm_size=2048,
        reduce_fsm_by=0.875,
        sampler=SamplerParams(k=100, alpha=1.0, log_pos=True),
    )
    assert alpha_regime(params) == "log_pos"


def test_sweep_membership_with_config():
    config = load_experiment_config(CONFIG_PATH)
    params = RunParams(
        seq_id="s",
        fsm_id="logexp_reduced_fsm_2048",
        fsm_size=2048,
        reduce_fsm_by=0.875,
        sampler=SamplerParams(k=100, alpha=1.0, log_pos=False),
    )
    assert "alpha" in sweep_membership(params, config)


def test_build_design_run_summary_dedupes():
    sols = [_make_result(1.0, 2.0, 2.1), _make_result(2.0, 1.0, 1.1, "002")]
    run = _make_run("seq1", 100, 1.0, True, 0.875, 2048, sols)
    config = load_experiment_config(CONFIG_PATH)
    summaries = build_design_run_summaries([run], config)
    assert len(summaries) == 1
    assert "alpha" in summaries[0].sweeps


def test_cross_sequence_aggregation():
    summaries = []
    for seq, hv in [("a", 0.4), ("b", 0.6)]:
        from pareto_designer.views.experiment_report.metrics import design_run_summary

        run = _make_run(seq, 100, 1.0, True, 0.875, 2048, [_make_result(1, 2, 2)])
        summaries.append(design_run_summary(run, hv, ["k"]))
    agg = aggregate_cross_sequence(summaries, "k")
    assert len(agg) == 1
    assert agg[0].n_seq == 2
    assert math.isclose(agg[0].hv_mean, 0.5)


def test_sort_design_runs():
    from pareto_designer.views.experiment_report.metrics import (
        design_run_summary,
        sort_design_runs,
    )

    runs = [
        design_run_summary(
            _make_run("b", 50, 1.0, True, 0.875, 2048, [_make_result(1, 2, 2)]),
            0.4,
            ["k"],
        ),
        design_run_summary(
            _make_run("a", 100, 1.0, True, 0.875, 4096, [_make_result(1, 2, 2)]),
            0.5,
            ["k"],
        ),
        design_run_summary(
            _make_run("a", 150, 1.0, True, 0.875, 2048, [_make_result(1, 2, 2)]),
            0.6,
            ["k"],
        ),
    ]
    ordered = sort_design_runs(runs)
    assert [r.seq_id for r in ordered] == ["a", "a", "b"]
    assert ordered[0].fsm_size == 4096
    assert ordered[1].k == 150


def test_correl_formula_in_sweep_sheet(tmp_path: Path):
    from openpyxl import Workbook
    from openpyxl.worksheet.formula import ArrayFormula

    from pareto_designer.views.experiment_report.excel_tables import (
        ALPHA_CORREL_SECTIONS,
        write_correlations_block,
        write_data_block,
        design_run_row,
    )
    from pareto_designer.views.experiment_report.metrics import design_run_summary

    wb = Workbook()
    ws = wb.active
    header_row = 3
    run_a = design_run_summary(
        _make_run("seq_a", 100, 0.0, False, 0.875, 2048, [_make_result(1, 2, 2)]),
        0.3,
        ["alpha"],
    )
    run_b = design_run_summary(
        _make_run("seq_b", 100, 1.0, False, 0.875, 2048, [_make_result(1, 2, 2)]),
        0.7,
        ["alpha"],
    )
    data = write_data_block(
        ws,
        header_row,
        1,
        DESIGN_RUN_HEADERS,
        (design_run_row(r) for r in [run_a, run_b]),
    )
    correl_col = len(DESIGN_RUN_HEADERS) + 2
    correl = write_correlations_block(
        ws,
        header_row,
        correl_col,
        data,
        ALPHA_CORREL_SECTIONS,
    )
    const_formula = ws.cell(row=header_row + 1, column=correl_col + 1).value
    assert isinstance(const_formula, ArrayFormula)
    assert const_formula.text.startswith("=CORREL(IF(")
    assert correl.end_row == header_row + len(ALPHA_CORREL_SECTIONS)


def test_fsm_err_correl_rows(tmp_path: Path):
    from openpyxl import Workbook

    from pareto_designer.views.experiment_report.excel_tables import FSM_CORREL_SECTIONS
    from pareto_designer.views.experiment_report.metrics import design_run_summary

    wb = Workbook()
    ws = wb.active
    header_row = 3
    run = design_run_summary(
        _make_run("seq_a", 100, 1.0, True, 0.875, 2048, [_make_result(1, 2, 2)]),
        0.5,
        ["fsm_size"],
    )
    data = write_data_block(
        ws,
        header_row,
        1,
        DESIGN_RUN_HEADERS,
        (design_run_row(run),),
    )
    correl_col = len(DESIGN_RUN_HEADERS) + 2
    write_correlations_block(ws, header_row, correl_col, data, FSM_CORREL_SECTIONS)
    size_formula = ws.cell(row=header_row + 1, column=correl_col + 1).value
    err_formula = ws.cell(row=header_row + 2, column=correl_col + 1).value
    assert size_formula.startswith("=CORREL(")
    assert "$C$" in size_formula
    assert err_formula.startswith("=CORREL(")
    assert "$O$" in err_formula


def test_grouped_bar_chart(tmp_path: Path):
    import zipfile

    from openpyxl import Workbook

    from pareto_designer.views.experiment_report.excel_charts import add_grouped_bar_chart

    wb = Workbook()
    ws = wb.active
    add_grouped_bar_chart(
        ws,
        "A1",
        title="test: hypervolume vs k",
        category_labels=["50", "100", "150"],
        series=[("seq_a", [0.1, 0.2, 0.12]), ("seq_b", [0.15, 0.18, 0.22])],
    )
    out = tmp_path / "chart.xlsx"
    wb.save(out)
    with zipfile.ZipFile(out) as zf:
        xml = zf.read("xl/charts/chart1.xml").decode()
    assert xml.count("<ser>") == 2
    assert "barChart" in xml
    assert "clustered" in xml or 'grouping val="clustered"' in xml
    assert "<legend>" in xml


@patch("pareto_designer.shared.seq_design_utils.run_grid.SequenceDesigner")
def test_run_design_grid_delegates(mock_designer_cls):
    mock_designer = MagicMock()
    mock_exporter = MagicMock()
    mock_exporter._results = []
    mock_designer.with_score_function_builder.return_value = mock_designer
    mock_designer.with_target_sequence.return_value = mock_designer
    mock_designer.with_fsm_context.return_value = mock_designer
    mock_designer.with_sampler.return_value = mock_designer
    mock_designer.run.return_value = mock_exporter
    mock_designer_cls.return_value = mock_designer

    score_builder = MagicMock()
    fsm_ctx = MagicMock()
    fsm_ctx.reduce_fsm_by = 0.875
    fsm_ctx.fsm_id = "logexp_reduced_fsm_2048"
    fsm_ctx.motif_id = "MA0267.1"
    fsm_builder = MagicMock()
    fsm_builder.iter_contexts.return_value = [fsm_ctx]

    seq_file = Path("seq.txt")
    batches = run_design_grid(
        [seq_file],
        score_builder,
        fsm_builder,
        k_values=[100],
        sampler_alpha=["1.0"],
        reduce_fsm_by=[0.875],
        mode=GridMode.RUN,
    )
    assert seq_file.stem in batches
    mock_designer.run.assert_called_with(dry_run=False)
