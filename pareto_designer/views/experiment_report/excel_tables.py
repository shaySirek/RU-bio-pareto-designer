from __future__ import annotations

from typing import Any, Iterable, Sequence

from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet

from pareto_designer.views.experiment_report.metrics import (
    filter_design_runs_by_sweep,
    sort_design_runs,
    sort_solutions,
)
from pareto_designer.views.experiment_report.excel_schema import (
    DESIGN_RUN_TABLE,
    OVERVIEW_CHECKLIST_TABLE,
    SOLUTION_TABLE,
    ExcelTableSpec,
    GroupedExcelTableSpec,
)
from pareto_designer.views.experiment_report.models import (
    DesignRunSummary,
    ExperimentConfig,
    RangeRef,
    SolutionRecord,
)

SEQ_BORDER = Border(top=Side(style="medium"))

SWEEP_SHEETS = (
    ("Sweep alpha", "alpha"),
    ("Sweep K", "k"),
    ("Sweep FSM size", "fsm_size"),
)


def write_section_title(ws: Worksheet, row: int, title: str, *, col: int = 1) -> int:
    ws.cell(row=row, column=col, value=title).font = Font(bold=True, size=12)
    return row + 2


def write_data_block(
    ws: Worksheet,
    row: int,
    col: int,
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
    *,
    keys: Sequence[str] | None = None,
    seq_border: bool = False,
) -> RangeRef:
    col_keys = keys if keys is not None else headers
    col_map = {key: col + idx for idx, key in enumerate(col_keys)}
    n_cols = len(headers)
    for idx, header in enumerate(headers):
        ws.cell(row=row, column=col + idx, value=header).font = Font(bold=True)
    data_start = row + 1
    end_row = data_start - 1
    prev_seq_id: str | None = None
    for record in rows:
        end_row += 1
        if seq_border and prev_seq_id is not None and record[0] != prev_seq_id:
            for c in range(col, col + n_cols):
                ws.cell(row=end_row, column=c).border = SEQ_BORDER
        for idx, _header in enumerate(headers):
            ws.cell(row=end_row, column=col + idx, value=_excel_cell_value(record[idx]))
        if seq_border:
            prev_seq_id = str(record[0])
    return RangeRef(
        start_row=data_start,
        end_row=end_row,
        col_map=col_map,
        sheet_name=ws.title,
    )


def _excel_cell_value(value: Any) -> Any:
    if isinstance(value, float):
        return value if value == value else None
    return value


def _write_column_groups(
    ws: Worksheet,
    row: int,
    col: int,
    spec: GroupedExcelTableSpec,
) -> None:
    col_idx = col + len(spec.leading)
    for group in spec.groups:
        start_col = col_idx
        col_idx += len(group.columns)
        cell = ws.cell(row=row, column=start_col, value=group.title)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        if len(group.columns) > 1:
            ws.merge_cells(
                start_row=row,
                start_column=start_col,
                end_row=row,
                end_column=col_idx - 1,
            )


def write_table(
    ws: Worksheet,
    row: int,
    col: int,
    spec: ExcelTableSpec | GroupedExcelTableSpec,
    records: Iterable[Any],
    *,
    seq_border: bool = False,
) -> RangeRef:
    header_row = row
    if isinstance(spec, GroupedExcelTableSpec):
        _write_column_groups(ws, row, col, spec)
        header_row = row + 1
    return write_data_block(
        ws,
        header_row,
        col,
        spec.headers,
        spec.rows(records),
        keys=spec.keys,
        seq_border=seq_border,
    )


DESIGN_RUN_HEADERS = DESIGN_RUN_TABLE.headers


def design_run_row(summary: DesignRunSummary) -> list[Any]:
    return list(DESIGN_RUN_TABLE.row(summary))


SOLUTION_HEADERS = SOLUTION_TABLE.headers


def solution_row(record: SolutionRecord) -> list[Any]:
    return list(SOLUTION_TABLE.row(record))


def freeze_header_row(ws: Worksheet, header_row: int) -> None:
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def write_overview_sheet(
    wb,
    config: ExperimentConfig | None,
    checklist: list[tuple[str, str, str, bool]],
) -> None:
    ws = wb.create_sheet("Overview", 0)
    row = write_section_title(ws, 1, "Experiment overview")
    if config:
        ws.cell(row=row, column=1, value="Config name")
        ws.cell(row=row, column=2, value=config.name)
        row += 1
        ws.cell(row=row, column=1, value="Results root")
        ws.cell(row=row, column=2, value=config.fixed.get("results_root", ""))
        row += 2
    row = write_section_title(ws, row, "Expected runs checklist")
    write_table(
        ws,
        row,
        1,
        OVERVIEW_CHECKLIST_TABLE,
        (
            {"seq_id": s, "sweep": sw, "metadata_path": p, "exists": e}
            for s, sw, p, e in checklist
        ),
    )


def write_summary_sheet(wb, design_runs: list[DesignRunSummary]) -> None:
    ws = wb.create_sheet("Summary")
    write_table(
        ws,
        1,
        1,
        DESIGN_RUN_TABLE,
        sort_design_runs(design_runs),
        seq_border=True,
    )


def write_sweep_sheet(
    wb,
    sheet_title: str,
    sweep: str,
    design_runs: list[DesignRunSummary],
) -> None:
    runs = sort_design_runs(filter_design_runs_by_sweep(design_runs, sweep))
    if not runs:
        return
    ws = wb.create_sheet(sheet_title)
    header_row = write_section_title(ws, 1, "Design runs")
    write_table(
        ws,
        header_row,
        1,
        DESIGN_RUN_TABLE,
        runs,
        seq_border=True,
    )


def write_sweep_sheets(wb, design_runs: list[DesignRunSummary]) -> None:
    for sheet_title, sweep in SWEEP_SHEETS:
        write_sweep_sheet(wb, sheet_title, sweep, design_runs)


def write_solutions_sheet(wb, solutions: list[SolutionRecord]) -> None:
    ws = wb.create_sheet("Solutions")
    write_table(
        ws,
        1,
        1,
        SOLUTION_TABLE,
        sort_solutions(solutions),
        seq_border=True,
    )
    freeze_header_row(ws, 1)
